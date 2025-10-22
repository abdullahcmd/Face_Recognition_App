import pickle
import face_recognition
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Load model.pkl
with open("../model.pkl", "rb") as f:
    data = pickle.load(f)

known_encodings = data["encodings"]
known_names = data["names"]

# 📘 Excel file path
ATTENDANCE_FILE = "attendance.xlsx"


def initialize_attendance_file():
    """Create Excel file with student names if it doesn't exist"""
    if not os.path.exists(ATTENDANCE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Headers
        ws.append(["Name", "Status", "Last Marked"])

        # Add all known students as 'Absent' initially
        for name in known_names:
            ws.append([name, "Absent", ""])

        wb.save(ATTENDANCE_FILE)
        print("✅ Created new attendance.xlsx file")


def mark_attendance(name):
    """Mark student as present if recognized"""
    initialize_attendance_file()  # Ensure file exists

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    if name in ("Unknown", "No face detected"):
        print(f"⚠️ No valid recognition for {name}, attendance not updated.")
        return

    # Find the name in Excel and mark present
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == name:
            row[1].value = "Present"
            row[2].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            found = True
            break

    if not found:
        # If name not found in the list, append new record
        ws.append([name, "Present", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    wb.save(ATTENDANCE_FILE)
    print(f"✅ Attendance marked for {name}")


def get_attendance():
    """Return all attendance records as a list of dicts"""
    initialize_attendance_file()  # Ensure file exists

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    attendance_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        attendance_list.append({
            "Name": row[0],
            "Status": row[1],
            "LastMarked": row[2]
        })

    return attendance_list


def recognize_face(image):
    """Return recognized name or 'Unknown'"""
    encs = face_recognition.face_encodings(image)
    if len(encs) == 0:
        return "No face detected"

    face_encoding = encs[0]
    matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.5)
    face_distances = face_recognition.face_distance(known_encodings, face_encoding)

    best_match_index = np.argmin(face_distances)
    if matches[best_match_index]:
        return known_names[best_match_index]
    else:
        return "Unknown"
